# apps/costing/views.py

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from collections import defaultdict
from django.db import models
from apps.products.models import Product
from apps.bom.models import BOMItem, WoodPart
import io
from decimal import Decimal
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


@login_required
def costing_list(request):
    """
    Shows all active products with optional search filtering.
    """
    search_query = request.GET.get('search', '').strip()

    products = Product.objects.filter(
        is_deleted=False,
        active=True
    ).order_by('product_name').prefetch_related(
        'bom_items__resource', 'wood_parts__resource'
    )

    if search_query:
        products = products.filter(
            models.Q(product_name__icontains=search_query) |
            models.Q(product_code__icontains=search_query)
        )

    portfolio_total = Decimal('0')
    for product in products:
        # Total Product Cost = Standard BOM + Dimensional BOM (see
        # CLAUDE.md for the history of this rule — it used to be
        # Standard BOM only).
        bom_cost = sum(
            (item.cost for item in product.bom_items.all()),
            Decimal('0')
        )
        dimensional_cost = sum(
            (part.cost for part in product.wood_parts.all()),
            Decimal('0')
        )
        product.total_cost = bom_cost + dimensional_cost
        portfolio_total += product.total_cost

    context = {
        'page_title': 'Cost Sheets',
        'products': products,
        'search_query': search_query,
        'portfolio_total': portfolio_total,
    }
    return render(request, 'costing/list.html', context)


@login_required
def cost_sheet(request, product_pk):
    """
    Displays the full cost sheet for a single product.

    Cost logic (rule change — see CLAUDE.md for history):
    - Standard BOM (BOMItem) and Dimensional BOM (WoodPart) are two
      distinct, non-overlapping cost buckets.
    - Grand total = Standard BOM total + Dimensional BOM total.
    - The category summary / pie chart merges both buckets per
      category, since a category (e.g. "Carpentry Materials") can
      have costs from both a BOMItem and a WoodPart line.
    """
    product = get_object_or_404(Product, pk=product_pk, is_deleted=False)

    bom_items = BOMItem.objects.filter(
        product=product
    ).select_related('resource').order_by(
        'resource__category', 'resource__resource_name'
    )

    wood_parts = WoodPart.objects.filter(
        product=product
    ).select_related('resource').order_by(
        'resource__category', 'part_name'
    )

    # ── Standard BOM: group by category ────────────────────────────
    bom_by_category = defaultdict(list)
    for item in bom_items:
        bom_by_category[item.resource.category].append(item)

    bom_category_totals = {}
    for category, items in bom_by_category.items():
        bom_category_totals[category] = sum(item.cost for item in items)

    # ── Dimensional BOM: group by category, now genuinely costed ───
    wood_by_category = defaultdict(list)
    for part in wood_parts:
        wood_by_category[part.resource.category].append(part)

    wood_category_totals = {}
    for category, parts in wood_by_category.items():
        wood_category_totals[category] = sum(part.cost for part in parts)

    # ── Totals ───────────────────────────────────────────────────────
    standard_bom_total    = sum(bom_category_totals.values())
    dimensional_bom_total = sum(wood_category_totals.values())
    grand_total            = standard_bom_total + dimensional_bom_total

    # ── Category summary: merge both buckets per category ──────────
    # A category can have cost from BOTH a BOMItem and a WoodPart
    # (e.g. "Carpentry Materials" might have hinges in Standard BOM
    # and Teak Wood in Dimensional BOM) — combine before computing
    # percentages so the pie chart reflects the true total.
    combined_category_totals = defaultdict(Decimal)
    for category, total in bom_category_totals.items():
        combined_category_totals[category] += total
    for category, total in wood_category_totals.items():
        combined_category_totals[category] += total

    category_summary = []
    for category, total in sorted(
        combined_category_totals.items(),
        key=lambda x: x[1],
        reverse=True
    ):
        percentage = (float(total) / float(grand_total) * 100) if grand_total > 0 else 0
        category_summary.append({
            'category':   category,
            'total':      total,
            'percentage': round(percentage, 1),
        })

    context = {
        'page_title': f'Cost Sheet — {product.product_code}',
        'product': product,

        # Standard BOM
        'bom_by_category':     dict(bom_by_category),
        'bom_category_totals': bom_category_totals,
        'standard_bom_total':  standard_bom_total,

        # Dimensional BOM
        'wood_by_category':     dict(wood_by_category),
        'wood_category_totals': wood_category_totals,
        'dimensional_bom_total': dimensional_bom_total,

        # Totals
        'grand_total':        grand_total,

        # Category summary (merged)
        'category_summary':   category_summary,
    }
    return render(request, 'costing/cost_sheet.html', context)


@login_required
def export_full_workbook(request):
    """
    Exports the entire database as a single Excel workbook with 7 sheets:
        1. Products
        2. Resources
        3. Suppliers
        4. SupplierRates
        5. BOM
        6. Dimensions
        7. CostSummary

    All data comes from the live database — not from any previously
    uploaded files. Nothing is saved to disk; the workbook is built
    in memory and sent directly to the browser as a download.
    """
    from apps.products.models import Product
    from apps.resources.models import Resource
    from apps.suppliers.models import Supplier, ResourceSupplier
    from apps.bom.models import BOMItem, WoodPart

    # ── Fetch all data up front ───────────────────────────────────────
    # select_related() tells Django to JOIN the related table in one
    # SQL query instead of hitting the database once per row.
    products = Product.objects.filter(
        is_deleted=False
    ).order_by('product_name')

    resources = Resource.objects.all().order_by('category', 'resource_name')

    suppliers = Supplier.objects.all().order_by('supplier_name')

    supplier_rates = ResourceSupplier.objects.select_related(
        'supplier', 'resource'
    ).order_by('resource__category', 'resource__resource_name')

    bom_items = BOMItem.objects.select_related(
        'product', 'resource'
    ).filter(
        product__is_deleted=False
    ).order_by('product__product_name', 'resource__category', 'resource__resource_name')

    wood_parts = WoodPart.objects.select_related(
        'product', 'resource'
    ).filter(
        product__is_deleted=False
    ).order_by('product__product_name', 'part_name')

    # ── Create workbook ───────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Shared style helpers ──────────────────────────────────────────
    # Define once, reuse across all 7 sheets.

    def make_header_style():
        return {
            'fill': PatternFill("solid", fgColor="1a1f3a"),
            'font': Font(color="FFFFFF", bold=True, size=10),
            'align': Alignment(horizontal="center", vertical="center"),
        }

    def write_headers(ws, headers):
        """Writes a styled header row to a sheet."""
        ws.append(headers)
        s = make_header_style()
        for cell in ws[ws.max_row]:
            cell.fill = s['fill']
            cell.font = s['font']
            cell.alignment = s['align']
        ws.row_dimensions[ws.max_row].height = 20

    def set_col_widths(ws, widths):
        """Sets column widths. widths is a list of integers, one per column."""
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def stripe_row(ws, row_num):
        """Light grey fill for even rows — makes large tables easier to read."""
        if row_num % 2 == 0:
            grey = PatternFill("solid", fgColor="F7F7F7")
            for cell in ws[row_num]:
                cell.fill = grey

    thin = Side(style='thin', color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_border(ws):
        """Applies a thin border to every cell that has content."""
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    # ════════════════════════════════════════════════════════════════
    # SHEET 1 — Products
    # ════════════════════════════════════════════════════════════════
    ws_products = wb.active
    ws_products.title = "Products"

    write_headers(ws_products, [
        "Product Code", "Product Name", "Active", "Created At"
    ])

    for i, p in enumerate(products, start=2):
        ws_products.append([
            p.product_code,
            p.product_name,
            "Yes" if p.active else "No",
            p.created_at.strftime('%d %b %Y') if p.created_at else '',
        ])
        stripe_row(ws_products, i)

    set_col_widths(ws_products, [18, 40, 8, 16])
    apply_border(ws_products)
    ws_products.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # SHEET 2 — Resources
    # ════════════════════════════════════════════════════════════════
    ws_resources = wb.create_sheet("Resources")

    write_headers(ws_resources, [
        "Resource Name", "Category", "Unit",
        "Master Rate (₹)", "Effective Rate (₹)", "Rate Source",
        "Override Rate (₹)", "Override Reason", "Active"
    ])

    for i, r in enumerate(resources, start=2):
        rate_source = r.effective_rate_source
        ws_resources.append([
            r.resource_name,
            r.category,
            r.unit,
            float(r.rate),
            float(r.effective_rate),
            rate_source['label'],
            float(r.manual_override_rate) if r.manual_override_rate else '',
            r.override_reason or '',
            "Yes" if r.active else "No",
        ])
        # Format the rate columns as currency
        row = ws_resources[ws_resources.max_row]
        row[3].number_format = '₹#,##0.00'
        row[4].number_format = '₹#,##0.00'
        if r.manual_override_rate:
            row[6].number_format = '₹#,##0.00'
        stripe_row(ws_resources, i)

    set_col_widths(ws_resources, [28, 18, 10, 16, 16, 18, 16, 24, 8])
    apply_border(ws_resources)
    ws_resources.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # SHEET 3 — Suppliers
    # ════════════════════════════════════════════════════════════════
    ws_suppliers = wb.create_sheet("Suppliers")

    write_headers(ws_suppliers, [
        "Supplier Name", "Phone Number", "GST Number",
        "Active", "Resources Linked"
    ])

    for i, s in enumerate(suppliers, start=2):
        ws_suppliers.append([
            s.supplier_name,
            s.phone_number or '',
            s.gst_number or '',
            "Yes" if s.active else "No",
            s.resource_count,
        ])
        stripe_row(ws_suppliers, i)

    set_col_widths(ws_suppliers, [32, 16, 18, 8, 16])
    apply_border(ws_suppliers)
    ws_suppliers.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # SHEET 4 — Supplier Rates
    # ════════════════════════════════════════════════════════════════
    ws_rates = wb.create_sheet("SupplierRates")

    write_headers(ws_rates, [
        "Supplier Name", "Resource Name", "Category", "Unit",
        "Supplier Rate (₹)", "Preferred", "Active"
    ])

    for i, link in enumerate(supplier_rates, start=2):
        ws_rates.append([
            link.supplier.supplier_name,
            link.resource.resource_name,
            link.resource.category,
            link.resource.unit,
            float(link.supplier_rate),
            "Yes" if link.preferred else "No",
            "Yes" if link.active else "No",
        ])
        row = ws_rates[ws_rates.max_row]
        row[4].number_format = '₹#,##0.00'

        # Highlight preferred supplier rows in soft green
        if link.preferred:
            green = PatternFill("solid", fgColor="E8F5E9")
            for cell in ws_rates[ws_rates.max_row]:
                cell.fill = green
        else:
            stripe_row(ws_rates, i)

    set_col_widths(ws_rates, [28, 28, 18, 10, 18, 10, 8])
    apply_border(ws_rates)
    ws_rates.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # SHEET 5 — BOM
    # ════════════════════════════════════════════════════════════════
    ws_bom = wb.create_sheet("BOM")

    write_headers(ws_bom, [
        "Product Code", "Product Name", "Resource Name",
        "Category", "Unit", "Quantity", "Rate (₹)", "Cost (₹)"
    ])

    for i, item in enumerate(bom_items, start=2):
        ws_bom.append([
            item.product.product_code,
            item.product.product_name,
            item.resource.resource_name,
            item.resource.category,
            item.resource.unit,
            float(item.quantity),
            float(item.resource.effective_rate),
            float(item.cost),
        ])
        row = ws_bom[ws_bom.max_row]
        row[5].number_format = '#,##0.0000'
        row[6].number_format = '₹#,##0.00'
        row[7].number_format = '₹#,##0.00'
        stripe_row(ws_bom, i)

    set_col_widths(ws_bom, [16, 32, 28, 18, 10, 12, 14, 14])
    apply_border(ws_bom)
    ws_bom.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # SHEET 6 — Dimensions (WoodParts)
    # ════════════════════════════════════════════════════════════════
    ws_dims = wb.create_sheet("Dimensions")

    write_headers(ws_dims, [
        "Product Code", "Product Name", "Part Name", "Resource",
        "Width", "Breadth", "Length", "Pieces", "Formula", "Calc. Qty",
        "Rate (₹)", "Cost (₹)"
    ])

    for i, part in enumerate(wood_parts, start=2):
        ws_dims.append([
            part.product.product_code,
            part.product.product_name,
            part.part_name,
            part.resource.resource_name,
            float(part.width),
            float(part.breadth),
            float(part.length),
            part.pieces,
            part.get_formula_type_display(),
            float(part.calculated_quantity),
            float(part.rate),
            float(part.cost),
        ])
        row = ws_dims[ws_dims.max_row]
        row[9].number_format = '#,##0.0000'
        row[10].number_format = '₹#,##0.00'
        row[11].number_format = '₹#,##0.00'
        stripe_row(ws_dims, i)

    set_col_widths(ws_dims, [16, 32, 24, 24, 10, 10, 10, 8, 14, 12, 14, 14])
    apply_border(ws_dims)
    ws_dims.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════
    # SHEET 7 — Cost Summary (one row per product)
    # ════════════════════════════════════════════════════════════════
    ws_summary = wb.create_sheet("CostSummary")

    write_headers(ws_summary, [
        "Product Code", "Product Name", "BOM Items", "Dimension Entries",
        "Standard BOM Cost (₹)", "Dimensional BOM Cost (₹)",
        "Total Cost (₹)", "Active"
    ])

    # Group BOM items AND wood parts by product to calculate
    # per-product totals. A product's Total Cost is now the sum of
    # both buckets (rule change — see CLAUDE.md for history).
    from collections import defaultdict
    product_costs = defaultdict(lambda: {
        'items': 0, 'bom_total': Decimal('0'),
        'dimensions': 0, 'dimensional_total': Decimal('0'),
    })

    for item in bom_items:
        pk = item.product.pk
        product_costs[pk]['product'] = item.product
        product_costs[pk]['items'] += 1
        product_costs[pk]['bom_total'] += item.cost

    for part in wood_parts:
        pk = part.product.pk
        product_costs[pk]['product'] = part.product
        product_costs[pk]['dimensions'] += 1
        product_costs[pk]['dimensional_total'] += part.cost

    # Also include products that have no BOM/dimension data yet
    # so every product appears in the summary.
    all_products_map = {p.pk: p for p in products}

    summary_rows = []
    for pk, p in all_products_map.items():
        data = product_costs.get(pk, {})
        bom_total = data.get('bom_total', Decimal('0'))
        dimensional_total = data.get('dimensional_total', Decimal('0'))
        summary_rows.append({
            'product':          p,
            'items':            data.get('items', 0),
            'dimensions':       data.get('dimensions', 0),
            'bom_total':        bom_total,
            'dimensional_total': dimensional_total,
            'total':            bom_total + dimensional_total,
        })

    # Sort by total cost descending so most expensive products are at the top
    summary_rows.sort(key=lambda x: x['total'], reverse=True)

    for i, row_data in enumerate(summary_rows, start=2):
        ws_summary.append([
            row_data['product'].product_code,
            row_data['product'].product_name,
            row_data['items'],
            row_data['dimensions'],
            float(row_data['bom_total']),
            float(row_data['dimensional_total']),
            float(row_data['total']),
            "Yes" if row_data['product'].active else "No",
        ])
        row = ws_summary[ws_summary.max_row]
        row[4].number_format = '₹#,##0.00'
        row[5].number_format = '₹#,##0.00'
        row[6].number_format = '₹#,##0.00'
        stripe_row(ws_summary, i)

    # Grand total row at the bottom of cost summary
    grand_total_all = sum(r['total'] for r in summary_rows)
    ws_summary.append([])
    ws_summary.append([
        "TOTAL", "",
        sum(r['items'] for r in summary_rows),
        sum(r['dimensions'] for r in summary_rows),
        float(sum(r['bom_total'] for r in summary_rows)),
        float(sum(r['dimensional_total'] for r in summary_rows)),
        float(grand_total_all),
        "",
    ])
    gt_row = ws_summary[ws_summary.max_row]
    grand_fill = PatternFill("solid", fgColor="1a1f3a")
    grand_font = Font(color="FFFFFF", bold=True)
    for cell in gt_row:
        cell.fill = grand_fill
        cell.font = grand_font
    gt_row[4].number_format = '₹#,##0.00'
    gt_row[5].number_format = '₹#,##0.00'
    gt_row[6].number_format = '₹#,##0.00'

    set_col_widths(ws_summary, [18, 38, 12, 16, 20, 22, 18, 8])
    apply_border(ws_summary)
    ws_summary.freeze_panes = "A2"

    # ── Save to buffer and return as download ─────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Product_Costing_Export_{date.today().strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response